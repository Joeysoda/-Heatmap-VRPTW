"""
Data loader for hospital AGV scheduling datasets.

This module loads:
- nodes from robot_node.xlsx
- edges from robot_edge.xlsx
- orders from robot_order.xlsx

It also normalizes:
- node coordinates via true node lookup
- hard time-window fields (three-stage mapping)
- charging station extraction
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
import os
import re
from typing import Dict, List, Optional, Sequence, Tuple
import xml.etree.ElementTree as ET
import zipfile

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not found. Using built-in XLSX reader.")

from .config import (
    DEFAULT_CHARGING_STATIONS,
    ROBOT_EDGE_FILE,
    ROBOT_NODE_FILE,
    ROBOT_ORDER_FILE,
)


NodeCoord = Tuple[float, float, int]

_XLSX_NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XLS_RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_COL_RE = re.compile(r"^([A-Z]+)")


def _excel_col_to_index(cell_ref: str) -> int:
    m = _COL_RE.match(cell_ref.upper())
    if not m:
        return 0
    col = m.group(1)
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return max(0, idx - 1)


def _decode_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    out: List[str] = []
    for si in root.findall("x:si", _XLSX_NS):
        parts = [t.text or "" for t in si.findall(".//x:t", _XLSX_NS)]
        out.append("".join(parts))
    return out


def _resolve_sheet_target(zf: zipfile.ZipFile, preferred_names: Sequence[str]) -> Optional[str]:
    if "xl/workbook.xml" not in zf.namelist() or "xl/_rels/workbook.xml.rels" not in zf.namelist():
        return None

    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rid_to_target: Dict[str, str] = {}
    for rel in rels.findall(f"{{{_XLS_RELS_NS}}}Relationship"):
        rel_type = rel.attrib.get("Type", "")
        if rel_type.endswith("/worksheet"):
            rid_to_target[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")

    preferred = {_normalize_header(name) for name in preferred_names}
    sheets = wb.find("x:sheets", _XLSX_NS)
    if sheets is None:
        return None

    chosen_target: Optional[str] = None
    for sheet in sheets.findall("x:sheet", _XLSX_NS):
        rid = sheet.attrib.get(f"{{{_REL_NS}}}id", "")
        target = rid_to_target.get(rid)
        if not target:
            continue
        if _normalize_header(sheet.attrib.get("name")) in preferred:
            chosen_target = target
            break
        if chosen_target is None:
            chosen_target = target

    if chosen_target is None:
        return None
    return f"xl/{chosen_target.lstrip('/')}"


def _coerce_xlsx_cell_value(cell_elem: ET.Element, shared_strings: Sequence[str]):
    cell_type = cell_elem.attrib.get("t", "")
    value_elem = cell_elem.find("x:v", _XLSX_NS)
    inline_elem = cell_elem.find("x:is/x:t", _XLSX_NS)

    if cell_type == "inlineStr":
        return inline_elem.text if inline_elem is not None else ""
    if value_elem is None:
        return None

    raw = value_elem.text
    if raw is None:
        return None

    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except Exception:
            return raw
    if cell_type == "b":
        return raw == "1"
    if cell_type in ("str", "e"):
        return raw

    try:
        num = float(raw)
    except ValueError:
        return raw
    if num.is_integer():
        return int(num)
    return num


def _iter_xlsx_rows_builtin(path: str, preferred_names: Sequence[str]):
    if not os.path.exists(path):
        return

    try:
        with zipfile.ZipFile(path) as zf:
            target = _resolve_sheet_target(zf, preferred_names)
            if not target or target not in zf.namelist():
                return
            shared_strings = _decode_shared_strings(zf)
            with zf.open(target) as sheet_stream:
                for _, row_elem in ET.iterparse(sheet_stream, events=("end",)):
                    if row_elem.tag != f"{{{_XLSX_NS['x']}}}row":
                        continue
                    values: Dict[int, object] = {}
                    max_idx = -1
                    for cell in row_elem.findall("x:c", _XLSX_NS):
                        cell_ref = cell.attrib.get("r", "A1")
                        col_idx = _excel_col_to_index(cell_ref)
                        values[col_idx] = _coerce_xlsx_cell_value(cell, shared_strings)
                        max_idx = max(max_idx, col_idx)

                    row = [None] * (max_idx + 1) if max_idx >= 0 else []
                    for col_idx, v in values.items():
                        row[col_idx] = v
                    yield tuple(row)
                    row_elem.clear()
    except Exception as exc:
        print(f"Error reading xlsx with built-in parser: {path} ({exc})")


def _iter_xlsx_rows(path: str, preferred_names: Sequence[str]):
    if HAS_OPENPYXL:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = _pick_sheet(wb, preferred_names)
        try:
            for row in ws.iter_rows(values_only=True):
                yield row
        finally:
            wb.close()
        return

    yield from _iter_xlsx_rows_builtin(path, preferred_names)


@dataclass
class _NodeRecord:
    node_code: str
    x: float
    y: float
    z: int
    map_code: str
    node_type: str


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _build_header_index(header_row: Sequence[object]) -> Dict[str, int]:
    index: Dict[str, int] = {}
    for i, name in enumerate(header_row):
        normalized = _normalize_header(name)
        if normalized and normalized not in index:
            index[normalized] = i
    return index


def _pick_sheet(workbook, preferred_names: Sequence[str]):
    preferred = {_normalize_header(name) for name in preferred_names}
    for name in workbook.sheetnames:
        if _normalize_header(name) in preferred:
            return workbook[name]
    return workbook.active


def _safe_float(value: object, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_datetime(value: object) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date (days since 1899-12-30, with fractional day time).
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(value))
        except Exception:
            return None
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        # Sometimes values arrive as numeric strings from xlsx cells.
        numeric = float(text)
        return datetime(1899, 12, 30) + timedelta(days=numeric)
    except ValueError:
        pass
    try:
        # Supports formats like "2025-03-01 10:00:00"
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def _extract_floor_from_map(map_code: object) -> int:
    """
    Derive an integer floor index from map_code.

    The source data has map codes like 021, 042, 0604, 0251.
    We keep a conservative parser so z stays stable and bounded.
    """
    if map_code is None:
        return 0

    text = str(map_code).strip().upper()
    if not text:
        return 0
    if "B" in text:
        return -1

    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return 0

    # Prefer first 1-2 meaningful digits (keeps values small).
    try:
        if len(digits) >= 2:
            floor = int(digits[:2])
        else:
            floor = int(digits)
    except ValueError:
        return 0

    if floor > 10:
        floor = floor // 10
    return max(-2, min(floor, 10))


def _get_node_lookup(nodes: Sequence[dict]) -> Dict[str, _NodeRecord]:
    lookup: Dict[str, _NodeRecord] = {}
    for node in nodes:
        code = str(node.get("node_code", "")).strip()
        if not code:
            continue
        lookup[code] = _NodeRecord(
            node_code=code,
            x=float(node.get("x", 0.0)),
            y=float(node.get("y", 0.0)),
            z=int(node.get("z", 0)),
            map_code=str(node.get("map_code", "")),
            node_type=str(node.get("node_type", "")).lower(),
        )
    return lookup


def _parse_node_coords_fallback(node_code: str) -> NodeCoord:
    """
    Fallback parser when node_code is missing in node table.
    Format examples: 0604$SITE-00063
    """
    if not node_code:
        return (0.0, 0.0, 0)
    parts = node_code.split("$")
    map_code = parts[0] if parts else ""
    site_code = parts[1] if len(parts) > 1 else ""
    floor = _extract_floor_from_map(map_code)

    site_digits = "".join(ch for ch in site_code if ch.isdigit())
    if site_digits:
        num = int(site_digits)
        return (float((num % 100) * 10), float((num // 100) * 10), floor)
    return (0.0, 0.0, floor)


def load_robot_nodes() -> List[dict]:
    """Load nodes from robot_node.xlsx, prioritizing sheet 'robot_node'."""
    if not os.path.exists(ROBOT_NODE_FILE):
        print(f"robot_node.xlsx not found: {ROBOT_NODE_FILE}")
        return _get_fallback_nodes()
    try:
        rows = _iter_xlsx_rows(ROBOT_NODE_FILE, ("robot_node", "sheet1"))
        header = next(rows, None)
        if header is None:
            return _get_fallback_nodes()
        idx = _build_header_index(header)

        code_i = idx.get("node_code", 0)
        name_i = idx.get("node_name")
        x_i = idx.get("node_x")
        y_i = idx.get("node_y")
        map_i = idx.get("map_code")
        type_i = idx.get("type_name")
        if type_i is None:
            type_i = idx.get("type_code")

        nodes: List[dict] = []
        for row in rows:
            node_code_raw = row[code_i] if code_i is not None and code_i < len(row) else None
            if node_code_raw is None:
                continue
            node_code = str(node_code_raw).strip()
            if not node_code:
                continue

            map_code = ""
            if map_i is not None and map_i < len(row) and row[map_i] is not None:
                map_code = str(row[map_i]).strip()

            node_type = ""
            if type_i is not None and type_i < len(row) and row[type_i] is not None:
                node_type = str(row[type_i]).strip().lower()

            node_name = ""
            if name_i is not None and name_i < len(row) and row[name_i] is not None:
                node_name = str(row[name_i]).strip()

            x = _safe_float(row[x_i] if x_i is not None and x_i < len(row) else None)
            y = _safe_float(row[y_i] if y_i is not None and y_i < len(row) else None)
            z = _extract_floor_from_map(map_code or node_code.split("$")[0])

            nodes.append(
                {
                    "node_code": node_code,
                    "x": x,
                    "y": y,
                    "z": z,
                    "map_code": map_code,
                    "node_type": node_type,
                    "node_name": node_name,
                }
            )

        if not nodes:
            return _get_fallback_nodes()
        print(f"Loaded {len(nodes)} nodes from robot_node.xlsx")
        return nodes
    except Exception as exc:
        print(f"Error loading robot_node.xlsx: {exc}")
        return _get_fallback_nodes()


def load_robot_edges() -> List[dict]:
    """Load edges from robot_edge.xlsx."""
    if not os.path.exists(ROBOT_EDGE_FILE):
        print(f"robot_edge.xlsx not found: {ROBOT_EDGE_FILE}")
        return []
    try:
        rows = _iter_xlsx_rows(ROBOT_EDGE_FILE, ("robot_edge", "sheet1"))
        header = next(rows, None)
        if header is None:
            return []
        idx = _build_header_index(header)

        edge_i = idx.get("edge_code", 0)
        s_i = idx.get("start_node", 1)
        e_i = idx.get("end_node", 2)
        len_i = idx.get("edge_length", 10)
        vel_i = idx.get("robot_velocity", 11)
        time_i = idx.get("taketime", 12)
        map_i = idx.get("map_code", 9)

        edges: List[dict] = []
        for row in rows:
            edge_code_raw = row[edge_i] if edge_i < len(row) else None
            if edge_code_raw is None:
                continue
            edge = {
                "edge_code": str(edge_code_raw),
                "start_node": str(row[s_i]) if s_i < len(row) and row[s_i] is not None else "",
                "end_node": str(row[e_i]) if e_i < len(row) and row[e_i] is not None else "",
                "length": _safe_float(row[len_i] if len_i < len(row) else None),
                "velocity": _safe_float(row[vel_i] if vel_i < len(row) else None, default=1.0),
                "taketime": _safe_float(row[time_i] if time_i < len(row) else None),
                "map_code": str(row[map_i]) if map_i < len(row) and row[map_i] is not None else "",
            }
            edges.append(edge)

        print(f"Loaded {len(edges)} edges from robot_edge.xlsx")
        return edges
    except Exception as exc:
        print(f"Error loading robot_edge.xlsx: {exc}")
        return []


def _monotonic_datetimes(values: List[Optional[datetime]]) -> List[Optional[datetime]]:
    out = values[:]
    for i in range(1, len(out)):
        prev = out[i - 1]
        curr = out[i]
        if prev is None:
            continue
        if curr is None:
            out[i] = prev
            continue
        if curr < prev:
            out[i] = prev
    return out


def _apply_time_windows(order: dict, reference_time: datetime) -> None:
    """
    Three-stage hard time-window mapping:
    pickup = [start_time, put_pre_time]
    delivery = [put_pre_time, get_pre_time]
    final deadline = finish_time
    """
    create_time = _safe_datetime(order.get("create_time"))
    start_time = _safe_datetime(order.get("start_time"))
    put_pre_time = _safe_datetime(order.get("put_pre_time"))
    get_pre_time = _safe_datetime(order.get("get_pre_time"))
    finish_time = _safe_datetime(order.get("finish_time"))

    pickup_start = start_time or create_time
    pickup_end = put_pre_time or pickup_start or get_pre_time or finish_time
    delivery_start = put_pre_time or pickup_end
    delivery_end = get_pre_time or finish_time or delivery_start
    final_deadline = finish_time or delivery_end

    timeline = _monotonic_datetimes(
        [pickup_start, pickup_end, delivery_start, delivery_end, final_deadline]
    )
    pickup_start, pickup_end, delivery_start, delivery_end, final_deadline = timeline

    valid = all(
        t is not None for t in (pickup_start, pickup_end, delivery_start, delivery_end, final_deadline)
    )

    order["pickup_tw_start"] = pickup_start
    order["pickup_tw_end"] = pickup_end
    order["delivery_tw_start"] = delivery_start
    order["delivery_tw_end"] = delivery_end
    order["final_deadline"] = final_deadline
    order["time_window_valid"] = bool(valid)

    if not valid:
        order["pickup_tw_start_sec"] = None
        order["pickup_tw_end_sec"] = None
        order["delivery_tw_start_sec"] = None
        order["delivery_tw_end_sec"] = None
        order["final_deadline_sec"] = None
        return

    def to_seconds(dt: datetime) -> float:
        return float((dt - reference_time).total_seconds())

    order["pickup_tw_start_sec"] = to_seconds(pickup_start)
    order["pickup_tw_end_sec"] = to_seconds(pickup_end)
    order["delivery_tw_start_sec"] = to_seconds(delivery_start)
    order["delivery_tw_end_sec"] = to_seconds(delivery_end)
    order["final_deadline_sec"] = to_seconds(final_deadline)


def load_robot_orders(limit: Optional[int] = None, node_lookup: Optional[Dict[str, _NodeRecord]] = None) -> List[dict]:
    """Load orders and map coordinates using true node lookup."""
    if not os.path.exists(ROBOT_ORDER_FILE):
        print(f"robot_order.xlsx not found: {ROBOT_ORDER_FILE}")
        return _get_fallback_orders()

    if node_lookup is None:
        node_lookup = _get_node_lookup(load_robot_nodes())

    try:
        rows = _iter_xlsx_rows(ROBOT_ORDER_FILE, ("robot_order", "sheet1"))
        header = next(rows, None)
        if header is None:
            return _get_fallback_orders()
        idx = _build_header_index(header)

        id_i = idx.get("order_no", 0)
        name_i = idx.get("order_name", 1)
        s_node_i = idx.get("start_node", 13)
        e_node_i = idx.get("end_node", 14)

        create_i = idx.get("create_time", 8)
        start_i = idx.get("start_time", 9)
        put_i = idx.get("put_pre_time", 10)
        get_i = idx.get("get_pre_time", 11)
        finish_i = idx.get("finish_time", 12)

        orders: List[dict] = []

        for row_idx, row in enumerate(rows, start=1):
            if limit is not None and len(orders) >= limit:
                break

            order_id_raw = row[id_i] if id_i < len(row) else None
            if order_id_raw is None:
                continue
            order_id = str(order_id_raw).strip()
            if not order_id:
                continue

            start_node = str(row[s_node_i]).strip() if s_node_i < len(row) and row[s_node_i] else ""
            end_node = str(row[e_node_i]).strip() if e_node_i < len(row) and row[e_node_i] else ""

            start_rec = node_lookup.get(start_node)
            end_rec = node_lookup.get(end_node)

            if start_rec is None:
                sx, sy, sz = _parse_node_coords_fallback(start_node)
            else:
                sx, sy, sz = start_rec.x, start_rec.y, start_rec.z

            if end_rec is None:
                ex, ey, ez = _parse_node_coords_fallback(end_node)
            else:
                ex, ey, ez = end_rec.x, end_rec.y, end_rec.z

            order = {
                "id": order_id,
                "order_name": str(row[name_i]).strip() if name_i < len(row) and row[name_i] else "",
                "start_node": start_node,
                "end_node": end_node,
                "start_x": float(sx),
                "start_y": float(sy),
                "start_z": int(sz),
                "end_x": float(ex),
                "end_y": float(ey),
                "end_z": int(ez),
                "create_time": row[create_i] if create_i < len(row) else None,
                "start_time": row[start_i] if start_i < len(row) else None,
                "put_pre_time": row[put_i] if put_i < len(row) else None,
                "get_pre_time": row[get_i] if get_i < len(row) else None,
                "finish_time": row[finish_i] if finish_i < len(row) else None,
                "source_row_index": row_idx,
            }
            orders.append(order)

        if not orders:
            return _get_fallback_orders()

        # Build reference time from available order timestamps.
        candidates: List[datetime] = []
        for order in orders:
            for key in ("create_time", "start_time", "put_pre_time", "get_pre_time", "finish_time"):
                dt = _safe_datetime(order.get(key))
                if dt is not None:
                    candidates.append(dt)
        reference_time = min(candidates) if candidates else datetime(1970, 1, 1)

        for order in orders:
            _apply_time_windows(order, reference_time)

        print(f"Loaded {len(orders)} orders from robot_order.xlsx")
        return orders
    except Exception as exc:
        print(f"Error loading robot_order.xlsx: {exc}")
        return _get_fallback_orders()


def extract_charging_stations(nodes: Sequence[dict]) -> List[NodeCoord]:
    """
    Extract charging stations from node attributes first, then node code heuristics.
    """
    charging_stations: List[NodeCoord] = []
    seen = set()
    charge_keywords = ("charge", "stopcharge", "charging", "dock", "充电")

    for node in nodes:
        node_type = str(node.get("node_type", "")).lower()
        node_code = str(node.get("node_code", "")).lower()
        node_name = str(node.get("node_name", "")).lower()
        is_charge = (
            any(k in node_type for k in charge_keywords)
            or any(k in node_code for k in charge_keywords)
            or any(k in node_name for k in charge_keywords)
        )
        if not is_charge:
            continue
        coord = (float(node.get("x", 0.0)), float(node.get("y", 0.0)), int(node.get("z", 0)))
        if coord in seen:
            continue
        charging_stations.append(coord)
        seen.add(coord)

    if not charging_stations:
        print("No charging stations found in node data, using default locations")
        charging_stations = list(DEFAULT_CHARGING_STATIONS)

    print(f"Found {len(charging_stations)} charging stations")
    return charging_stations


def _get_fallback_nodes() -> List[dict]:
    return [
        {"node_code": "NODE_1", "node_name": "NODE_1", "x": 0.0, "y": 0.0, "z": 0, "map_code": "0604", "node_type": ""},
        {"node_code": "NODE_2", "node_name": "NODE_2", "x": 100.0, "y": 0.0, "z": 0, "map_code": "0604", "node_type": ""},
        {"node_code": "NODE_3", "node_name": "NODE_3", "x": 0.0, "y": 100.0, "z": 1, "map_code": "042", "node_type": ""},
        {"node_code": "NODE_4", "node_name": "NODE_4", "x": 100.0, "y": 100.0, "z": 1, "map_code": "042", "node_type": ""},
    ]


def _get_fallback_orders() -> List[dict]:
    now = datetime.now()
    return [
        {
            "id": "ORDER_1",
            "order_name": "Fallback Order 1",
            "start_x": 0.0,
            "start_y": 0.0,
            "start_z": 0,
            "end_x": 100.0,
            "end_y": 100.0,
            "end_z": 1,
            "start_node": "NODE_1",
            "end_node": "NODE_4",
            "pickup_tw_start": now,
            "pickup_tw_end": now,
            "delivery_tw_start": now,
            "delivery_tw_end": now,
            "final_deadline": now,
            "pickup_tw_start_sec": 0.0,
            "pickup_tw_end_sec": 0.0,
            "delivery_tw_start_sec": 0.0,
            "delivery_tw_end_sec": 0.0,
            "final_deadline_sec": 0.0,
            "time_window_valid": True,
        }
    ]


def load_all_data(order_limit: Optional[int] = 50):
    """Load nodes, edges, orders, and charging stations."""
    print("\n" + "=" * 60)
    print("Loading Hospital Robot Data")
    print("=" * 60)

    nodes = load_robot_nodes()
    node_lookup = _get_node_lookup(nodes)
    edges = load_robot_edges()
    orders = load_robot_orders(limit=order_limit, node_lookup=node_lookup)
    charging_stations = extract_charging_stations(nodes)

    print("=" * 60)
    print("Data loading complete:")
    print(f"  - Nodes: {len(nodes)}")
    print(f"  - Edges: {len(edges)}")
    print(f"  - Orders: {len(orders)}")
    print(f"  - Charging Stations: {len(charging_stations)}")
    print("=" * 60 + "\n")
    return nodes, edges, orders, charging_stations


if __name__ == "__main__":
    n, e, o, c = load_all_data(order_limit=10)
    print("Sample Node:", n[0] if n else None)
    print("Sample Edge:", e[0] if e else None)
    print("Sample Order:", o[0] if o else None)
    print("Charging Stations:", c[:3])

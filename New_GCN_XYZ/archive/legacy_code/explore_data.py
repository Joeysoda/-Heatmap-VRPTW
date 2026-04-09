"""
Data Exploration Script for Robot Data
Explores the structure of robot_data files to understand time window and charging constraints
"""

import openpyxl
import os

# Data directory - use absolute path
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(script_dir, "..", "..", ".."))
data_dir = os.path.join(project_root, "hospital-main", "robot_data")

print(f"Script directory: {script_dir}")
print(f"Project root: {project_root}")
print(f"Data directory: {data_dir}")
print(f"Data directory exists: {os.path.exists(data_dir)}")

# Files to explore
files = [
    "robot_order.xlsx",
    "robot_node.xlsx",
    "robot_map.xlsx",
    "robot_mission.xlsx",
    "robot_edge.xlsx",
    "robot_node_lift.xlsx"
]

print("=" * 80)
print("ROBOT DATA EXPLORATION")
print("=" * 80)

for file in files:
    filepath = os.path.join(data_dir, file)
    print(f"\n{'='*80}")
    print(f"FILE: {file}")
    print(f"{'='*80}")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Get dimensions
        max_row = min(ws.max_row, 11)  # Read header + 10 rows
        max_col = ws.max_column
        
        print(f"\nTotal rows in file: {ws.max_row}")
        print(f"Total columns: {max_col}")
        
        # Read header (first row)
        headers = []
        for col in range(1, max_col + 1):
            cell_value = ws.cell(1, col).value
            headers.append(str(cell_value) if cell_value is not None else f"Col{col}")
        
        print(f"\nColumns ({len(headers)}): {headers}")
        
        # Read first 5 data rows
        print(f"\nFirst 5 data rows:")
        for row_idx in range(2, min(7, max_row + 1)):
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row_idx, col).value
                row_data.append(cell_value)
            print(f"Row {row_idx-1}: {row_data}")
        
        # Check for time-related columns
        time_cols = [col for col in headers if any(keyword in str(col).lower() 
                     for keyword in ['time', 'window', 'tw', 'earliest', 'latest', 'deadline', 'release', 'arrive', 'depart', 'service'])]
        if time_cols:
            print(f"\n⏰ TIME-RELATED COLUMNS FOUND: {time_cols}")
        
        # Check for charging-related columns
        charge_cols = [col for col in headers if any(keyword in str(col).lower() 
                       for keyword in ['charge', 'battery', 'energy', 'capacity', 'station', 'power'])]
        if charge_cols:
            print(f"\n🔋 CHARGING-RELATED COLUMNS FOUND: {charge_cols}")
        
        # Check for coordinate columns
        coord_cols = [col for col in headers if any(keyword in str(col).lower() 
                      for keyword in ['x', 'y', 'z', 'coord', 'position', 'floor', 'node'])]
        if coord_cols:
            print(f"\n📍 COORDINATE COLUMNS FOUND: {coord_cols}")
        
        wb.close()
            
    except Exception as e:
        print(f"Error reading {file}: {e}")
        import traceback
        traceback.print_exc()

print("\n" + "=" * 80)
print("EXPLORATION COMPLETE")
print("=" * 80)

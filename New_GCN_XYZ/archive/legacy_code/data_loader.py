
import pandas as pd
import os
import numpy as np
from datetime import datetime
import openpyxl

def load_nodes(node_file_path):
    """
    Load node information from csv file.
    Returns a dictionary: {NodeName: (X, Y, Z)}
    """
    if not os.path.exists(node_file_path):
        raise FileNotFoundError(f"Node file not found: {node_file_path}")
        
    try:
        df = pd.read_csv(node_file_path)
        # Clean column names
        df.columns = [c.strip() for c in df.columns]
        
        node_map = {}
        required = {'NodeName', 'X', 'Y', 'Z'}
        if not required.issubset(set(df.columns)):
            raise ValueError("node.csv missing required columns: NodeName, X, Y, Z")
            
        for _, row in df.iterrows():
            node_map[str(row['NodeName']).strip()] = (float(row['X']), float(row['Y']), float(row['Z']))
                
        print(f"Loaded {len(node_map)} nodes from {node_file_path}")
        return node_map
    except Exception as e:
        print(f"Error loading nodes: {e}")
        return {}

def load_orders(instance_file_path, node_map=None):
    """
    Load orders from an xlsx file.
    Supports standard format or simplified 3-column format [No, StartNode, EndNode].
    Returns a list of dictionaries with coordinates.
    """
    if not os.path.exists(instance_file_path):
        raise FileNotFoundError(f"Instance file not found: {instance_file_path}")
        
    try:
        # Attempt to read the file
        df = pd.read_excel(instance_file_path)
    except Exception as e:
        print(f"Failed to read excel {instance_file_path}: {e}")
        return []
        
    orders = []
    
    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]
    col_map = {c.lower(): c for c in df.columns}
    
    # Logic to handle 3-column format (No, StartNode, EndNode) without headers or with different headers
    # This logic was refined in previous iterations to handle specific user files
    if not any(k in col_map for k in ['startx', 'starty', 'startz', 'start_node', 'startnode']):
        try:
            df_raw = pd.read_excel(instance_file_path, header=None)
            # Check if it looks like the 3-column format
            if df_raw.shape[1] >= 3:
                first_row = [str(x).strip().lower() for x in df_raw.iloc[0, :3].tolist()]
                # Check for header keywords
                if {'no', 'start_node', 'end_node'}.issubset(set(first_row)):
                    df = df_raw.iloc[1:].copy()
                    df.columns = ['no', 'start_node', 'end_node']
                    col_map = {c.lower(): c for c in df.columns}
        except Exception:
            pass

    for idx, row in df.iterrows():
        order_data = {}
        
        # 1. Parse ID
        if 'no' in col_map:
            order_data['id'] = row[col_map['no']]
        elif 'ordernum' in col_map:
            order_data['id'] = row[col_map['ordernum']]
        else:
            order_data['id'] = idx
            
        # 2. Parse Coordinates
        # Case A: Explicit coordinates
        if all(k in col_map for k in ['startx', 'starty', 'startz', 'endx', 'endy', 'endz']):
            order_data['start_x'] = float(row[col_map['startx']])
            order_data['start_y'] = float(row[col_map['starty']])
            order_data['start_z'] = float(row[col_map['startz']])
            order_data['end_x'] = float(row[col_map['endx']])
            order_data['end_y'] = float(row[col_map['endy']])
            order_data['end_z'] = float(row[col_map['endz']])
            
        # Case B: Node names (requires node_map)
        elif node_map and ('startnode' in col_map or 'start_node' in col_map) and ('endnode' in col_map or 'end_node' in col_map):
            s_key = 'startnode' if 'startnode' in col_map else 'start_node'
            e_key = 'endnode' if 'endnode' in col_map else 'end_node'
            
            s_node = str(row[col_map[s_key]]).strip()
            e_node = str(row[col_map[e_key]]).strip()
            
            if s_node in node_map and e_node in node_map:
                sx, sy, sz = node_map[s_node]
                ex, ey, ez = node_map[e_node]
                order_data['start_x'] = sx
                order_data['start_y'] = sy
                order_data['start_z'] = sz
                order_data['end_x'] = ex
                order_data['end_y'] = ey
                order_data['end_z'] = ez
            else:
                print(f"Warning: Node lookup failed for order {order_data['id']} ({s_node} -> {e_node})")
                continue
        
        # Case C: Fallback for unnamed columns (assuming 3-column format)
        else:
             try:
                vals = row.values
                if len(vals) >= 3 and node_map:
                     # Assume [ID, StartNode, EndNode]
                     s_node = str(vals[1]).strip()
                     e_node = str(vals[2]).strip()
                     if s_node in node_map and e_node in node_map:
                        sx, sy, sz = node_map[s_node]
                        ex, ey, ez = node_map[e_node]
                        order_data['start_x'] = sx
                        order_data['start_y'] = sy
                        order_data['start_z'] = sz
                        order_data['end_x'] = ex
                        order_data['end_y'] = ey
                        order_data['end_z'] = ez
                        order_data['id'] = vals[0]
                     else:
                         continue
                else:
                    continue
             except:
                 continue

        orders.append(order_data)
        
        print(f"Loaded {len(orders)} orders from {instance_file_path}")
    return orders


def load_orders_with_time_windows(order_file_path, node_map=None):
    """
    Load orders with time window constraints from robot_order.xlsx.
    Returns a list of dictionaries with order details including time windows.
    """
    if not os.path.exists(order_file_path):
        raise FileNotFoundError(f"Order file not found: {order_file_path}")
    
    try:
        df = pd.read_excel(order_file_path)
        df.columns = [str(c).strip() for c in df.columns]
        
        orders = []
        for idx, row in df.iterrows():
            order_data = {
                'id': row['order_no'],
                'order_name': row.get('order_name', ''),
                'start_node': row['start_node'],
                'end_node': row['end_node'],
                'start_address': row.get('start_address_name', ''),
                'end_address': row.get('end_address_name', ''),
                
                # Time windows
                'create_time': row['create_time'],
                'start_time': row.get('start_time'),
                'pickup_tw_start': row['create_time'],
                'pickup_tw_end': row['put_pre_time'],
                'delivery_tw_start': row['put_pre_time'],
                'delivery_tw_end': row['get_pre_time'],
                'finish_time': row.get('finish_time'),
                
                # Distance and time
                'distance': row.get('takelength', 0),
                'estimated_time': row.get('taketime', 0),
                
                # Robot assignment (if available)
                'robot_code': row.get('robot_code'),
            }
            
            # If node_map is provided, get coordinates
            if node_map:
                start_node_key = str(row['start_node']).strip()
                end_node_key = str(row['end_node']).strip()
                
                if start_node_key in node_map:
                    sx, sy, sz = node_map[start_node_key]
                    order_data['start_x'] = sx
                    order_data['start_y'] = sy
                    order_data['start_z'] = sz
                
                if end_node_key in node_map:
                    ex, ey, ez = node_map[end_node_key]
                    order_data['end_x'] = ex
                    order_data['end_y'] = ey
                    order_data['end_z'] = ez
            
            orders.append(order_data)
        
        print(f"Loaded {len(orders)} orders with time windows from {order_file_path}")
        return orders
        
    except Exception as e:
        print(f"Error loading orders with time windows: {e}")
        import traceback
        traceback.print_exc()
        return []


def extract_charging_stations(mission_file_path):
    """
    Extract charging station locations from robot_mission.xlsx.
    Returns a set of charging station node IDs and statistics.
    """
    if not os.path.exists(mission_file_path):
        raise FileNotFoundError(f"Mission file not found: {mission_file_path}")
    
    try:
        df = pd.read_excel(mission_file_path)
        df.columns = [str(c).strip() for c in df.columns]
        
        # Filter charging missions
        charging_missions = df[df['mission_type'] == 'charge']
        
        # Extract charging station nodes (end_node of charging missions)
        charging_stations = set()
        charging_times = []
        
        for idx, row in charging_missions.iterrows():
            station_node = str(row['end_node']).strip()
            charging_stations.add(station_node)
            
            # Collect charging time statistics
            if pd.notna(row.get('act_taketime')):
                charging_times.append(float(row['act_taketime']))
        
        # Calculate statistics
        avg_charging_time = np.mean(charging_times) if charging_times else 300.0
        min_charging_time = np.min(charging_times) if charging_times else 200.0
        max_charging_time = np.max(charging_times) if charging_times else 400.0
        
        charging_info = {
            'stations': list(charging_stations),
            'num_stations': len(charging_stations),
            'avg_charging_time': avg_charging_time,
            'min_charging_time': min_charging_time,
            'max_charging_time': max_charging_time,
            'total_charging_missions': len(charging_missions)
        }
        
        print(f"Extracted {len(charging_stations)} charging stations:")
        print(f"  Stations: {list(charging_stations)[:5]}...")
        print(f"  Avg charging time: {avg_charging_time:.2f}s")
        print(f"  Total charging missions: {len(charging_missions)}")
        
        return charging_info
        
    except Exception as e:
        print(f"Error extracting charging stations: {e}")
        import traceback
        traceback.print_exc()
        return {'stations': [], 'num_stations': 0, 'avg_charging_time': 300.0}


def load_edge_travel_times(edge_file_path):
    """
    Load edge travel times from robot_edge.xlsx.
    Returns a dictionary: {(start_node, end_node): travel_time}
    """
    if not os.path.exists(edge_file_path):
        raise FileNotFoundError(f"Edge file not found: {edge_file_path}")
    
    try:
        df = pd.read_excel(edge_file_path)
        df.columns = [str(c).strip() for c in df.columns]
        
        edge_times = {}
        for idx, row in df.iterrows():
            start_node = str(row['start_node']).strip()
            end_node = str(row['end_node']).strip()
            travel_time = float(row['taketime'])
            
            # Store both directions
            edge_times[(start_node, end_node)] = travel_time
            
        print(f"Loaded {len(edge_times)} edge travel times from {edge_file_path}")
        return edge_times
        
    except Exception as e:
        print(f"Error loading edge travel times: {e}")
        import traceback
        traceback.print_exc()
        return {}


def load_robot_nodes_from_xlsx(node_file_path):
    """
    Load node coordinates from robot_node.xlsx (handling the data format issue).
    Returns a dictionary: {node_id: (x, y, z)}
    """
    if not os.path.exists(node_file_path):
        raise FileNotFoundError(f"Node file not found: {node_file_path}")
    
    try:
        # Use openpyxl to handle the file
        wb = openpyxl.load_workbook(node_file_path, read_only=True, data_only=True)
        ws = wb.active
        
        node_map = {}
        # Skip first row (appears to be data, not header)
        for row_idx in range(1, min(ws.max_row + 1, 10000)):  # Limit to first 10k rows
            row_data = []
            for col in range(1, 6):  # 5 columns
                cell_value = ws.cell(row_idx, col).value
                row_data.append(cell_value)
            
            # Check if row has valid data
            if all(v is not None for v in row_data[:5]):
                try:
                    # Assuming format: [x_small, y_small, x_large, y_large, z]
                    x = float(row_data[2])  # Use x_large
                    y = float(row_data[3])  # Use y_large
                    z = float(row_data[4])  # z (floor)
                    
                    # Generate node_id (we'll need to match this with actual node IDs later)
                    node_id = f"NODE_{row_idx}"
                    node_map[node_id] = (x, y, z)
                except (ValueError, TypeError):
                    continue
        
        wb.close()
        print(f"Loaded {len(node_map)} nodes from {node_file_path}")
        return node_map
        
    except Exception as e:
        print(f"Error loading robot nodes: {e}")
        import traceback
        traceback.print_exc()
        return {}
